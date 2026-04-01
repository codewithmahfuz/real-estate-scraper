#!/usr/bin/env python3
# pip install requests beautifulsoup4 openpyxl
# If your system blocks pip (PEP 668): python3 -m venv .venv && .venv/bin/pip install requests beautifulsoup4 openpyxl
"""
Real estate style scraper — demo target: books.toscrape.com (legal practice site).
Maps book listings to property-style fields and exports to Excel.
"""

import sys
import time
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration — default scrape target (no configuration required)
# ---------------------------------------------------------------------------
BASE_URL = "https://books.toscrape.com"
START_PATH = "/catalogue/page-1.html"
OUTPUT_FILE = "real_estate_data.xlsx"
REQUEST_TIMEOUT = 30
PAGE_DELAY_SECONDS = 1

# Map star-rating CSS class names to numeric scores (One=1 … Five=5)
STAR_WORD_TO_NUMBER = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5,
}

# HTTP session reused for connection pooling and consistent headers
SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": "RealEstateDemoScraper/1.0 (+https://books.toscrape.com practice)",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9",
    }
)


# ---------------------------------------------------------------------------
# Helpers — rating and status normalization
# ---------------------------------------------------------------------------
def star_class_to_rating(star_rating_el):
    """Convert <p class='star-rating Three'> to integer 1–5; missing → None."""
    try:
        if star_rating_el is None:
            return None
        classes = star_rating_el.get("class") or []
        for name in classes:
            if name in STAR_WORD_TO_NUMBER:
                return STAR_WORD_TO_NUMBER[name]
    except (TypeError, AttributeError):
        pass
    return None


def availability_to_status(availability_el):
    """Map stock text to 'In Stock' / 'Out of Stock'."""
    try:
        if availability_el is None:
            return "Unknown"
        text = " ".join(availability_el.stripped_strings)
        lower = text.lower()
        if "out of stock" in lower:
            return "Out of Stock"
        if "in stock" in lower:
            return "In Stock"
        return text.strip() or "Unknown"
    except (AttributeError, TypeError):
        return "Unknown"


# ---------------------------------------------------------------------------
# Core scraping — fetch one page and parse listing cards
# ---------------------------------------------------------------------------
def parse_listings_from_html(html, page_url):
    """
    Parse all product cards on a catalogue page.
    Returns a list of dicts with property-style keys.
    """
    rows = []
    try:
        soup = BeautifulSoup(html, "html.parser")
        articles = soup.select("article.product_pod")
    except Exception as exc:
        print(f"Warning: could not parse HTML for {page_url}: {exc}", file=sys.stderr)
        return rows

    if not articles:
        try:
            print(f"Warning: no listings found on {page_url} (empty or unexpected layout).", file=sys.stderr)
        except OSError:
            pass
        return rows

    for art in articles:
        try:
            title_a = art.select_one("h3 a")
            price_el = art.select_one("p.price_color")
            star_el = art.select_one("p.star-rating")
            avail_el = art.select_one("p.instock.availability")

            name = (title_a.get("title") or "").strip() if title_a else ""
            if not name and title_a:
                name = title_a.get_text(strip=True)

            price_text = price_el.get_text(strip=True) if price_el else ""
            rating = star_class_to_rating(star_el)
            status = availability_to_status(avail_el)

            if not name and not price_text:
                continue

            rows.append(
                {
                    "Property Name": name or "(no title)",
                    "Property Price": price_text or "",
                    "Property Rating": rating if rating is not None else "",
                    "Property Status": status,
                }
            )
        except Exception as exc:
            print(f"Warning: skipped one listing on {page_url}: {exc}", file=sys.stderr)
            continue

    return rows


def fetch_page(url):
    """
    GET a page with full error handling.
    Returns (ok: bool, status_code: int|None, html: str|None, error_message: str|None).
    """
    try:
        response = SESSION.get(url, timeout=REQUEST_TIMEOUT)
    except requests.exceptions.ConnectionError as exc:
        return False, None, None, f"Connection error: {exc}"
    except requests.exceptions.Timeout as exc:
        return False, None, None, f"Request timed out: {exc}"
    except requests.exceptions.RequestException as exc:
        return False, None, None, f"Request failed: {exc}"

    code = response.status_code
    try:
        if code == 404:
            return False, code, None, "Page not found (404)."
        if code != 200:
            return False, code, None, f"Unexpected HTTP status: {code}"
        # books.toscrape.com serves UTF-8; requests may default to ISO-8859-1 — fix for correct £ etc.
        try:
            response.encoding = response.apparent_encoding or response.encoding or "utf-8"
        except (AttributeError, TypeError):
            response.encoding = "utf-8"
        html = response.text
        if not html or not html.strip():
            return False, code, None, "Empty response body."
        return True, code, html, None
    except Exception as exc:
        return False, code, None, f"Error while handling response: {exc}"


# ---------------------------------------------------------------------------
# Pagination — walk every catalogue page until no “next” link
# ---------------------------------------------------------------------------
def scrape_all_pages():
    """
    Scrape every paginated catalogue page; polite delay between requests.
    Returns (all_rows: list, pages_scraped: int, error_messages: list).
    """
    all_rows = []
    errors = []
    current_url = urljoin(BASE_URL + "/", START_PATH.lstrip("/"))
    page_index = 0

    while current_url:
        page_index += 1
        try:
            print(f"Scraping page {page_index}...")
        except OSError:
            pass

        ok, status, html, err = fetch_page(current_url)
        if not ok:
            errors.append(f"{current_url}: {err}")
            try:
                print(f"Error: {err}", file=sys.stderr)
            except OSError:
                pass
            break

        try:
            page_rows = parse_listings_from_html(html, current_url)
        except Exception as exc:
            errors.append(f"{current_url}: parse error: {exc}")
            page_rows = []

        if not page_rows:
            try:
                msg = f"No data extracted from {current_url}; stopping pagination."
                print(msg, file=sys.stderr)
                errors.append(msg)
            except OSError:
                pass
            break

        all_rows.extend(page_rows)

        try:
            soup = BeautifulSoup(html, "html.parser")
            next_a = soup.select_one("li.next a")
            if next_a and next_a.get("href"):
                current_url = urljoin(current_url, next_a["href"])
            else:
                current_url = None
        except Exception as exc:
            errors.append(f"Pagination error at {current_url}: {exc}")
            current_url = None

        if current_url:
            try:
                time.sleep(PAGE_DELAY_SECONDS)
            except Exception:
                pass

    return all_rows, page_index, errors


# ---------------------------------------------------------------------------
# Excel export — workbook, headers, bold/colored header row, column widths
# ---------------------------------------------------------------------------
def autosize_columns(ws, min_width=12, max_width=60):
    """Set column widths from cell content (rough auto-fit)."""
    try:
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = min_width
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, max_width)
    except Exception as exc:
        print(f"Warning: column auto-width failed: {exc}", file=sys.stderr)


def save_to_excel(rows, filepath):
    """Write rows to XLSX with professional header styling."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Listings"

        headers = ["Property Name", "Property Price", "Property Rating", "Property Status"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        body_align = Alignment(vertical="top", wrap_text=True)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, key in enumerate(headers, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
                c.alignment = body_align

        autosize_columns(ws)
        wb.save(filepath)
        return True, None
    except Exception as exc:
        return False, str(exc)


# ---------------------------------------------------------------------------
# Entry point — orchestrate scrape and export with top-level error handling
# ---------------------------------------------------------------------------
def main():
    """Run full pipeline: scrape all pages, save Excel, print summary."""
    try:
        rows, pages_done, scrape_errors = scrape_all_pages()
    except Exception as exc:
        print(f"Fatal error during scraping: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        if not rows:
            print("No properties collected. Excel file not written (empty data).", file=sys.stderr)
            if scrape_errors:
                for e in scrape_errors:
                    print(f"  — {e}", file=sys.stderr)
            sys.exit(2)
    except Exception:
        pass

    try:
        ok, err = save_to_excel(rows, OUTPUT_FILE)
        if not ok:
            print(f"Failed to save Excel: {err}", file=sys.stderr)
            sys.exit(3)
    except Exception as exc:
        print(f"Failed to save Excel: {exc}", file=sys.stderr)
        sys.exit(3)

    try:
        total = len(rows)
        print(f"Total {total} properties saved!")
    except OSError:
        pass


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(130)
